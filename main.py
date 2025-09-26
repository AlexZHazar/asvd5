# pip install pandas openpyxl PySide6
# pip install pywin32
# pip install PyInstaller
# pip install --upgrade PyInstaller pyinstaller-hooks-contrib

# pyinstaller --windowed --onefile --name ASVD_repairs main.py --icon=bin/OilTrain.ico

import ctypes
import subprocess
import os
import time
import win32gui
import win32con
import re
import sqlite3
import shutil

from glob import glob
from datetime import date
# from pprint import pprint
from datetime import datetime
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

import pandas as pd

from PySide6.QtWidgets import (QWidget, QFileDialog, QApplication, QPushButton, QLineEdit, QCheckBox, QVBoxLayout,
                               QLabel, QFormLayout, QComboBox, QDateEdit, QHBoxLayout, QSizePolicy, QLayout,
                               QMessageBox, QFileSystemModel, QDialog, QTreeView, QDialogButtonBox, QScrollArea,
                               QStyledItemDelegate, QCalendarWidget)
from PySide6.QtUiTools import QUiLoader
from PySide6.QtCore import QFile, QIODevice, QDate, Qt, QDir, QTimer, QUrl
from PySide6.QtGui import QFont, QIcon, QDesktopServices

import sqlite_objects as ss
from sqlite_interaction import sqlite_interaction as si
from sqlite_analysis import sqlite_analysis as sa

app_icon_png = "bin/OilTrain.png"

class XLSXLoader(QWidget):

    source_folder_path = ''
    settings_db = "bin/db_settings.sqlite"
    today = str(date.today())
    carriage_info_updated = False
    db_loading_log = ('',0)
    sa = None
    info_mark = '<span style="color: green;">INFO</span>'
    warning_mark = '<span style="color: blue;">WARNING</span>'
    error_mark = '<span style="color: red;">ERROR</span>'
    test_mark = '<span style="color: orange;">TEST</span>'
    base_type = {'work': 'рабочая', 'empty': 'пустая'}
    current_base_type_key = ''


    def __init__(self):
        super().__init__()

        self.read_config_file_for_dev_only()

        with sqlite3.connect(self.settings_db) as self.settings_conn:
            self.settings_cursor = self.settings_conn.cursor()

        self.carriage_info_formLayout = None

        loader = QUiLoader()
        ui_file = QFile("bin/xlsxloader.ui")
        ui_file.open(QIODevice.ReadOnly)  # type: ignore

        self.ui = loader.load(ui_file, self)
        ui_file.close()

        self.setWindowIcon(QIcon(app_icon_png))

        # Создаем layout с отступами и добавляем загруженный UI
        layout = QVBoxLayout()
        layout.setContentsMargins(10, 10, 10, 10)  # Установи нужные отступы
        layout.addWidget(self.ui)
        self.setLayout(layout)

        # # Подключение виджетов as specific objects
        # tab1
        self.folder_input: QLineEdit = self.ui.findChild(QLineEdit, "folderInput")
        self.select_folder_button: QPushButton = self.ui.findChild(QPushButton, "selectFolderButton")
        self.sqlite_db_input: QLineEdit = self.ui.findChild(QLineEdit, "sqliteDbInput")
        self.select_sqlite_db_button: QPushButton = self.ui.findChild(QPushButton, "selectSqliteDbButton")
        self.load_button: QPushButton = self.ui.findChild(QPushButton, "loadButton")
        self.refresh_checkbox: QCheckBox = self.ui.findChild(QCheckBox, "refreshCheckBox")
        self.log_scroll: QScrollArea = self.findChild(QScrollArea, "logScroll")
        self.log_text: QLabel = self.findChild(QLabel, "logText")

        self.checkBox_full_analysis: QCheckBox = self.ui.findChild(QCheckBox, "checkBox_full_analysis")
        self.reports_reports: QPushButton = self.ui.findChild(QPushButton, "pushButton_export_reports")
        # self.refreshCheckBox.isChecked()

        # tab2
        self.carriage_input: QLineEdit = self.ui.findChild(QLineEdit, "carriageInput")
        self.carriage_button: QPushButton = self.ui.findChild(QPushButton, "carriageButton")
        self.carriage_update_button: QPushButton = self.ui.findChild(QPushButton, "carriageUpdateButton")

        # tab3
        self.factory_input: QLineEdit = self.ui.findChild(QLineEdit, "factoryInput")
        self.factory_button: QPushButton = self.ui.findChild(QPushButton, "factoryButton")
        self.factory_name_edit: QLineEdit = self.ui.findChild(QLineEdit, "factoryEdit")
        self.factory_update_button: QPushButton = self.ui.findChild(QPushButton, "factoryUpdateButton")

        self.sqlite_db_input.setText(self._get_settings('sqlite_db_path', self.settings_cursor))
        self.sqlite_db_connect(self.sqlite_db_input.text())

        # Подключение сигналов
        self.carriage_button.clicked.connect(lambda: self.carriage_info(self.carriage_input.text()))
        self.carriage_update_button.clicked.connect(self.update_carriage_info)

        self.factory_button.clicked.connect(lambda: self.factory_info(self.factory_input.text()))
        self.factory_update_button.clicked.connect(self.update_factory_info)

        self.select_sqlite_db_button.clicked.connect(self.open_sqlite_db_file_dialog)

        # self.load_button.pressed.connect(lambda: self._loging_start_process(f"[{self.info_mark}] Загрузка и анализ запущены..."))
        # self.load_button.clicked.connect(lambda: QTimer.singleShot(2000, self.load_xlsx_files))  # запуск с задержкой
        # self.pushButton_export_reports.pressed.connect(lambda: self._loging_start_process(f"[{self.info_mark}] Формирование отчётов запущено..."))
        # self.pushButton_export_reports.clicked.connect(lambda: QTimer.singleShot(2000, self.export_reports))

        self.load_button.clicked.connect(lambda: self.clicked_button(f"[{self.info_mark}] Загрузка и анализ данных...", self.load_xlsx_files))
        self.reports_reports.clicked.connect(lambda: self.clicked_button(f"[{self.info_mark}] Формирование отчётов запущено...", self.export_reports))

    @staticmethod
    def stub(text):
        print('----')
        print(text)
        print('----')

    @staticmethod
    def read_config_file_for_dev_only(filename="sqlite_db_connection_dev.txt"):
        settings = {}
        # Получаем абсолютный путь относительно текущего исполняемого скрипта
        base_dir = os.path.abspath(os.path.dirname(__file__))
        file_path = os.path.join(base_dir, filename)

        if not os.path.exists(file_path):
            # print(f"Файл {filename} не найден в {base_dir}")
            return

        with open(file_path, "r", encoding="utf-8") as file:
            for line in file:
                line = line.strip()
                if not line or "=" not in line:
                    continue  # Пропустить пустые строки и некорректные форматы
                key, value = line.split("=", 1)
                settings[key.strip()] = value.strip()

        for k, v in settings.items():
            settings[k] = os.path.join(os.path.abspath(os.path.dirname(__file__)), v)
        # print(settings)

        loop_dict = ((str(settings.get("db_settings_path")), 'sqlite_db_path', settings.get('sqlite_db_path').replace('\\', '/')),
                     (str(settings.get("sqlite_db_path")), 'source_folder_path', settings.get('source_folder_path').replace('\\', '/')),
                     (str(settings.get("sqlite_db_path")).replace('asvd.db', 'empty.db'), 'source_folder_path', settings.get('source_folder_path').replace('\\', '/')),
                     )
        # pprint(loop_dict)

        for i in loop_dict:
            try:
                conn = sqlite3.connect(i[0])
                print(f"Успешное подключение к базе данных {i[0]}")
                conn.cursor().execute(f"update settings set value = '{i[2]}' where key = '{i[1]}'")
                conn.commit()
                conn.close()
                print(f"Соединение {i[0]} закрыто")
            except sqlite3.Error as e:
                print(f"Ошибка подключения: {e}")

    def clicked_button(self, msg, action):
        self.disable_elements(True, True)
        self._loging_start_process(msg)
        QTimer.singleShot(2000, action)  # запуск с задержкой

    def sqlite_db_connect(self, path):

        if os.path.isfile(path):
            abs_path = os.path.abspath(path)
            abs_path_text = abs_path.replace('\\', '/')

            try:
                with sqlite3.connect(abs_path) as self.sqlite_conn:
                    self.sqlite_conn.enable_load_extension(True)
                    self.sqlite_conn.load_extension("./bin/regexp")
                    self.sqlite_cursor = self.sqlite_conn.cursor()

                    self.sqlite_cursor.execute("select value from settings where key = 'data_storage'")
                    base_type_key = self.sqlite_cursor.fetchall()[0][0]
                    try:
                        if not bool(base_type_key):
                            raise
                        else:
                            message_t = (f"[{self.info_mark}] {datetime.now().strftime('%H:%M:%S')}",
                                         f"[{self.info_mark}] Подключена база данных: <b>{abs_path_text}</b>",
                                         f"[{self.info_mark}] Тип базы данных: <b>{self.base_type[base_type_key]}</b>")
                            self._logging_connection(message_t)

                            self.current_base_type_key = base_type_key

                    except:
                        self._logging_connection((f"[{self.error_mark}] Неверная база данных!",))
                        self.disable_elements(True)
                        raise

                    self.sqlite_cursor.execute("select value from settings where key = 'full_analysis'")

                    self.si = si(self.sqlite_cursor)

                    self.sqlite_cursor.execute(ss.all_views)

                    view_names = [row[0] for row in self.sqlite_cursor.fetchall()]
                    for d_view in view_names:
                        self.sqlite_cursor.execute(f"DROP VIEW IF EXISTS {d_view}")

                    l_views = [i for i in dir(ss) if i.startswith("view_")]
                    for v in l_views:
                        # print(v)
                        self.sqlite_cursor.execute(getattr(ss, v))

                    l_tables = [i for i in dir(ss) if i.startswith("table_")]
                    for t in l_tables:
                        # print(t)
                        try:
                            self.sqlite_cursor.execute(getattr(ss, t))
                        except Exception as e:
                            print(e)
                            raise

                    self.folder_input.setText(self._get_settings('source_folder_path', self.sqlite_cursor))
                    self.select_folder_button.clicked.connect(lambda: self.open_folder_dialog('source_folder_path', self.folder_input, 'Выберите папку с .xlsx файлами'))
                    self.disable_elements(False)

            except Exception as e:
                self._logging_connection((f"[{self.error_mark}] Ошибка подключения к базе данных: {abs_path}",))
        else:
            self._logging_connection((f"[{self.error_mark}] Путь не найден. Подключите базу данных",))
            self.disable_elements(True)

    def disable_elements(self, flag, db_elem=False):
        self.sqlite_db_input.setReadOnly(db_elem)
        self.select_sqlite_db_button.setDisabled(db_elem)
        self.folder_input.setReadOnly(flag)
        self.select_folder_button.setDisabled(flag)
        self.carriage_input.setReadOnly(flag)
        self.carriage_button.setDisabled(flag)
        self.carriage_update_button.setDisabled(flag)
        self.factory_input.setReadOnly(flag)
        self.factory_button.setDisabled(flag)
        self.factory_update_button.setDisabled(flag)
        self.load_button.setDisabled(flag)
        self.reports_reports.setDisabled(flag)
        # self.refresh_checkbox.setDisabled(flag)  # TODO: refreshCheckBox
        self.checkBox_full_analysis.setDisabled(flag)
        self.factory_name_edit.setDisabled(flag)

    def _logging_connection(self, messege_tuple: tuple):
        self._logging_message("*" * 130)
        for i in messege_tuple:
            self._logging_message(i)
        self._logging_message("=" * 40)
        self._logging_message("")

    def _loging_start_process(self, message: str):
        self._logging_message("*" * 65)
        self._logging_message(f"[{self.info_mark}] {datetime.now().strftime('%H:%M:%S')}")
        self._logging_message(message)
        self._logging_message("-" * 40)


    @staticmethod
    def create_xlsx_from_xltx(templete_name):

        now = datetime.now()
        short_name = re.sub(r'\.[^.]+$', '', templete_name)
        result_file_path = f"Reports/{short_name}_{now.strftime('%Y%m%d%H%M%S')}.xlsx"

        template_wb = load_workbook(templete_name)

        # Создаём новую рабочую книгу
        new_wb = Workbook()
        # Удаляем пустой лист, созданный по умолчанию
        new_wb.remove(new_wb.active)

        # Копируем все листы из шаблона
        for template_ws in template_wb.worksheets:
            new_ws = new_wb.create_sheet(template_ws.title)

            # Копируем значения и стили
            for row in template_ws.iter_rows():
                for cell in row:
                    new_cell = new_ws[cell.coordinate]
                    new_cell.value = cell.value

                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = copy(cell.number_format)
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)

            # Копируем ширину колонок
            for col_letter, dim in template_ws.column_dimensions.items():
                new_ws.column_dimensions[col_letter].width = dim.width

            # Копируем высоту строк (опционально)
            for row_idx, dim in template_ws.row_dimensions.items():
                new_ws.row_dimensions[row_idx].height = dim.height

            # Копируем объединённые ячейки
            for merged_range in template_ws.merged_cells.ranges:
                new_ws.merge_cells(str(merged_range))

        # # (по желанию) редактируем новый файл
        # new_wb["Sheet1"]["A1"] = "Создан из шаблона"

        # Сохраняем как обычный файл
        new_wb.save(result_file_path)

        return result_file_path

    def show_message(self, title, message):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle(title)
        msg.setText(message)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.setWindowIcon(QIcon(app_icon_png))
        msg.exec()
        self.disable_elements(False)

    def open_folder_dialog(self, key_dir, input_string, dlg_name):
        if not os.path.isdir(input_string.text()):
            default_dir = self._get_settings('default_' + key_dir, self.sqlite_cursor)
        else:
            default_dir = input_string.text()

        dlg = FolderOnlyPicker(title=dlg_name)
        dlg.set_default_path(default_dir)  # <-- Путь по умолчанию

        if dlg.exec():
            # print("Выбрана папка:", dlg.get_selected_folder())
            input_string.setText(dlg.get_selected_folder())
            self._set_settings(key_dir, input_string.text(), self.sqlite_cursor)
        else:
            pass
            # print("Отмена")

    def open_sqlite_db_file_dialog(self):

        if not os.path.isfile(self.sqlite_db_input.text()):
            # db_path = self._get_settings('default_sqlite_db_path', self.settings_cursor)
            db_folder_path = re.sub(r'/[^/]+$', '', self._get_settings('default_sqlite_db_path', self.settings_cursor))
        else:
            # db_path = self.sqlite_db_input.text()
            db_folder_path = re.sub(r'/[^/]+$', '', self.sqlite_db_input.text())

        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выберите файл базы данных", db_folder_path, "SQLite DB Files (*.db)")

        if file_path:
            self.sqlite_db_input.setText(file_path)
            self._set_settings('sqlite_db_path', self.sqlite_db_input.text(), self.settings_cursor)
            self.sqlite_db_connect(file_path)
            self.folder_input.setText(self._get_settings('source_folder_path', self.sqlite_cursor))

    def _logging_message(self, p_log_message=''):
        self.log_text.setText(self.log_text.text() + p_log_message+"<br>")
        QTimer.singleShot(500, lambda: self.log_scroll.verticalScrollBar().setValue(
            self.log_scroll.verticalScrollBar().maximum()
        ))

    def _get_settings(self, p_key, cursor):
        # print(p_key)
        cursor.execute(f"SELECT value FROM settings where key = '{p_key}'")
        return cursor.fetchall()[0][0]

    def _set_settings(self, p_key, p_value, cursor):
        script = 'select 1 from settings where key = ?'
        cursor.execute(script, (p_key,))
        if cursor.fetchone() is None:
            script = f"insert into settings (key, value) values ('{p_key}', '{p_value}')"
            cursor.execute(script)
        else:
            script = f"update settings set value = '{p_value}' where key = '{p_key}'"
            cursor.execute(script)
        cursor.execute('COMMIT')
        # print(script)

    def load_xlsx_files(self):
        empty_source = False
        all_dataframes = []
        self.source_folder_path = self.folder_input.text()

        self._logging_message(f"[{self.info_mark}] Загрузка файлов из:  {self.source_folder_path}<br>"+"-" * 40)

        if not os.path.isdir(self.source_folder_path):
            self._logging_message(f"[{self.error_mark}] Папка не найдена: {self.source_folder_path}")
            return

        xlsx_files = glob(os.path.join(self.source_folder_path, "*.xlsx"))
        if not xlsx_files:
            self._logging_message(f"[{self.warning_mark}] Файлы .xlsx не найдены")
            self._logging_message("=" * 40)
            empty_source = True

        if not empty_source:
            for file in xlsx_files:
                try:
                    df = pd.read_excel(file)
                    try:
                        df.columns = ["A", "B", "C", "D", "E"]  # Переименовываем 5 колонок
                    except Exception as e:
                        df.columns = ["A", "B", "C", "D"]  # Переименовываем 4 колонки
                    df["source_file"] = os.path.basename(file)  # �� добавляем колонку с именем файла
                    df["id"] = df.index + 1  # добавляем колонку id с автоинкрементом
                    all_dataframes.append(df)
                    self._logging_message(f"[{self.info_mark}] '{os.path.basename(file)}'")
                except Exception as e:
                    self._logging_message(f"[{self.error_mark}] '{os.path.basename(file)}': {str(e)}")

            if all_dataframes:
                full_df = pd.concat(all_dataframes, ignore_index=True)

                full_df = full_df.astype({
                    'A': 'string',
                    'B': 'string',
                    'C': 'string',
                    'D': 'string',
                    'E': 'string'
                })

                full_df.to_sql("tmp_xlsx_data", self.sqlite_conn, if_exists="replace", index=False)

                self.db_loading_log = self.si.sqlite_checking_process()

            else:
                self._logging_message(f"[{self.warning_mark}] Ни один файл не загружен")

            df_input_data_import = pd.read_sql("SELECT * FROM input_data_import_py", self.sqlite_conn)
            # print(df_input_data_import.head())

            rows = []
            for i, f in enumerate(df_input_data_import.source_file):

                df_file_eq = pd.read_sql(f"SELECT param, value FROM tmp_equipment where source_file = '{f}'", self.sqlite_conn)
                # print(df_file_eq.head())
                df_file_eq_transposed = df_file_eq.set_index("param").T
                # print(df_file_eq_transposed)

                df_file_eq_generic = pd.read_sql(
                    f"""SELECT * FROM input_data_import_py where source_file = '{f}'; """, self.sqlite_conn)

                df_file_eq_merged = pd.concat([df_file_eq_generic.reset_index(drop=True),
                                               df_file_eq_transposed.reset_index(drop=True)], axis=1)
                df_file_eq_merged["source_file"] = f
                # print(df_file_eq_merged)
                rows.append(df_file_eq_merged)
                # df_file_eq_merged.to_excel(f"output_{f}", index=False)
                # df_file_eq_generic.to_excel(f"output1_{f}", index=False)
                # df_file_eq_transposed.to_excel(f"output2_{f}", index=False)

            try:
                df_final = pd.concat(rows, ignore_index=True)
            except:
                # self._logging_message("-" * 40)
                self._logging_message(self.db_loading_log[0])
                self._logging_message("-" * 40)
                # self._logging_message(f"[{self.error_mark}] Ошибка: нет данных для обработки")
                self._logging_message(f"[{self.warning_mark}] Нет данных для обработки")
                self._logging_message("-" * 40)
                self._logging_message(f"[{self.info_mark}] Файлы перемещены в архив: /Archive/{self.today}/")
                self._logging_message("=" * 40)
                self._move_xlsx_files()
                self.disable_elements(False)
                raise
            # df_final.to_excel(f"output.xlsx", index=False)
            df_final.to_sql("inquiry_data_py", self.sqlite_conn, if_exists="replace", index=False)

            self.db_loading_log = self.si.sqlite_loading_process()

            self._logging_message(self.db_loading_log[0])

            if self.db_loading_log[1] == 0:
                self._logging_message("-" * 40)
                self._logging_message(f"[{self.info_mark}] Файлы перемещены в архив: /Archive/{self.today}/")
                self._logging_message("-" * 40)
                self._logging_message(f"[{self.info_mark}] Процесс загрузки завершён")
                self._logging_message("=" * 40)

                self._move_xlsx_files()
            else:
                self._logging_message("-" * 40)
                self._logging_message(f"[{self.error_mark}] Процесс завершился с ошибкой")
                self._logging_message(f"[{self.info_mark}] {datetime.now().strftime('%H:%M:%S')}")
                self._logging_message("=" * 40)

        try:
            self.sqlite_conn.commit()
        except:
            pass
        self.sqlite_conn.execute("VACUUM")
        self.replacement_analysis()
        self.show_message("Информация", "Операция выполнена")

    def _move_xlsx_files(self):
        archive_folder = os.path.join(self.source_folder_path, "Archive")

        archive_folder_date = os.path.join(archive_folder, self.today)
        # Проверяем, существует ли папка Archive, если нет - создаем
        if not os.path.exists(archive_folder_date):
            os.makedirs(archive_folder_date)

        # Получаем список всех файлов в исходной папке
        for file_name in os.listdir(self.source_folder_path):
            # Проверяем, что это файл с расширением .xlsx
            if file_name.endswith('.xlsx'):
                # Полный путь к исходному файлу
                source_file = os.path.join(self.source_folder_path, file_name)
                # Полный путь к файлу в папке Archive
                archive_file = os.path.join(archive_folder_date, file_name)

                # Перемещаем файл в папку Archive
                shutil.move(source_file, archive_file)
                # print(f"Файл {file_name} перемещен в {archive_folder_date}")

    def carriage_info(self, carriage_number):
        # int(carriage_number)
        try:
            int(carriage_number.strip())
            df = pd.read_sql_query(f"SELECT * FROM carriage where carriage_code = {carriage_number}", self.sqlite_conn)
            # print(df)
            # row_kv = df.iloc[0].T.reset_index()
            # row_kv.columns = ['key', 'value']
            record = df.iloc[0].to_dict()
            row_kv = pd.DataFrame({
                'key': list(record.keys()),
                'value': [str(v) if pd.notna(v) else '' for v in record.values()]  # Преобразуем всё в строки
            })
            # Добавляем счётчик строк
            row_kv.insert(0, 'row_number', range(1, len(row_kv) + 1))
            # print(row_kv)
            row_kv.to_sql('tmp_carriage_info', self.sqlite_conn, if_exists='replace', index=False)
            df = pd.read_sql_query("select * from tmp_carriage_info_v", self.sqlite_conn)
            self.carriage_info_formLayout = self.ui.findChild(QFormLayout, "formLayout")
            self.fill_form_layout(self.carriage_info_formLayout, df)
            self.connect_signals(self.carriage_info_formLayout)
        except ValueError as e:
            print(f'incorrect input value {carriage_number}\n{e}')

    def factory_info(self, factory_number):
        try:
            int(factory_number.strip())
            self.sqlite_cursor.execute(f"SELECT factory_name FROM factory where cast(factory_code as integer) = {factory_number}")
            factory_name = self.sqlite_cursor.fetchall()[0][0]
            # print(factory_name)
        except (ValueError, IndexError) as e:
            print(f'incorrect input value {factory_number}\n{e}')
            factory_name = ''
        self.factory_name_edit.setText(factory_name)

    def fill_form_layout(self, layout: QFormLayout, df):
        # Очистка текущего layout'а
        self.clear_layout(layout)

        # print(df)
        # Заполняем из DataFrame
        for idx, row in df.iterrows():
            key_ = row['key']
            text_name_ = row['key_text']
            value_ = row['value']
            map_key_ = row['map_key']
            editable_ = row['editable']
            boolean_ = row['boolean']
            list_ = row['list']
            date_ = row['date']

            label = QLabel(map_key_)
            label.setObjectName(f"{key_}_label")
            label_x = QLabel(' ')
            font_x = QFont("Courier New")
            label_x.setFont(font_x)
            label_x.setObjectName(f"{key_}_label_x")
            label_x.setStyleSheet("color: red;")
            # noinspection PyUnresolvedReferences
            label_x.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

            h_layout = QHBoxLayout()
            h_layout.setObjectName(f"{key_}_hblo")

            # print(text_name_)

            if date_ == 1:

                # print(key_)
                # print("'" + value_ + "'")

                dateEdit = QDateEdit()
                dateEdit.setObjectName(text_name_)
                date_value = QDate.fromString(value_, 'dd.MM.yyyy')
                if date_value.isValid():
                    dateEdit.setDate(date_value)
                else:
                    date_value = QDate.fromString('01.01.1900', 'dd.MM.yyyy')
                    dateEdit.setToolTip('01.01.1900 - означает, что дата отсутствует')
                    dateEdit.setDate(date_value)
                dateEdit.setCalendarPopup(True)

                h_layout.addWidget(label_x)
                h_layout.addWidget(dateEdit)
                layout.addRow(label, h_layout)

            elif list_ == 1:
                comboBox = QComboBox()
                comboBox.setObjectName(text_name_)
                # print(key_)

                df_list = pd.read_sql_query(f"select value from list_info where key = '{key_}'", self.sqlite_conn)
                comboBox.addItems(df_list["value"].astype(str).tolist())
                comboBox.setEditable(bool(editable_))
                # print(value_)
                # print(key_)

                if str(value_) in df_list["value"].values:
                    comboBox.setCurrentText(str(value_))
                else:
                    comboBox.setCurrentText(None)
                # print(value_)

                h_layout.addWidget(label_x)
                h_layout.addWidget(comboBox)
                layout.addRow(label, h_layout)

            elif boolean_ == 0:
                lineEdit = QLineEdit()
                lineEdit.setObjectName(text_name_)
                lineEdit.setText(str(value_))
                lineEdit.setReadOnly(not bool(editable_))

                h_layout.addWidget(label_x)
                h_layout.addWidget(lineEdit)
                layout.addRow(label, h_layout)

            else:
                checkbox = QCheckBox()
                checkbox.setObjectName(text_name_)
                checkbox.setChecked(bool(int(value_)))

                h_layout.addWidget(label_x)
                h_layout.addWidget(checkbox)
                layout.addRow(label, h_layout)

    def connect_signals(self, form_layout):

        for i in range(form_layout.rowCount()):
            # print(i)
            # noinspection PyUnresolvedReferences
            h_layout_widget = form_layout.itemAt(i, QFormLayout.FieldRole).layout()
            if isinstance(h_layout_widget, QHBoxLayout):
                for j in range(h_layout_widget.count()):
                    # print(j)
                    widget = h_layout_widget.itemAt(j).widget()
                    try:
                        name = widget.objectName()
                        # print(name)
                        if isinstance(widget, QLineEdit):
                            widget.textChanged.connect(lambda _, name_=name, hlw_=h_layout_widget: self.on_any_change(name_, hlw_))
                        elif isinstance(widget, QCheckBox):
                            widget.stateChanged.connect(lambda _, name_=name, hlw_=h_layout_widget: self.on_any_change(name_, hlw_))
                        elif isinstance(widget, QComboBox):
                            widget.currentTextChanged.connect(lambda _, name_=name, hlw_=h_layout_widget: self.on_any_change(name_, hlw_))
                        elif isinstance(widget, QDateEdit):
                            widget.dateChanged.connect(lambda _, name_=name, hlw_=h_layout_widget: self.on_any_change(name_, hlw_))
                    except AttributeError as e:
                        print(f'Ошибка:\n{e}')

    def on_any_change(self, *args):
        h_layout = args[1]
        lable_x_name = args[0].replace('text', 'label_x')
        # print(h_layout, lable_x_name)

        for i in range(h_layout.count()):
            item = h_layout.itemAt(i)
            widget = item.widget()
            # print(widget)
            # print(lable_x_name)
            if isinstance(widget, QLabel) and widget.objectName() == lable_x_name:
                widget.setText('*')
                self.carriage_info_updated = True
                break

        # print("Что-то изменилось!")

    def clear_layout(self, layout):

        while layout.rowCount():
            layout.removeRow(0)

    @staticmethod
    def extract_widget_data(widget_or_layout):
        results = []

        def recurse(obj):
            # Если это layout — обходим все его элементы
            if isinstance(obj, QLayout):
                for i in range(obj.count()):
                    item = obj.itemAt(i)

                    if item.widget():
                        recurse(item.widget())
                    elif item.layout():
                        recurse(item.layout())

            # Если это QWidget, проверим тип
            elif isinstance(obj, QWidget):
                name = obj.objectName()
                if not name:
                    return  # Пропускаем, если имя пустое

                if isinstance(obj, QLineEdit):
                    results.append((name, obj.text()))
                elif isinstance(obj, QCheckBox):
                    results.append((name, obj.isChecked()))
                elif isinstance(obj, QComboBox):
                    results.append((name, obj.currentText()))
                elif isinstance(obj, QDateEdit):
                    if obj.date() > QDate(1900, 1, 1):
                        results.append((name, obj.date().toString("dd.MM.yyyy")))
                    else:
                        results.append((name, None))
                else:
                    # Рекурсивно обходим его layout, если есть
                    if obj.layout():
                        recurse(obj.layout())

        recurse(widget_or_layout)

        # Преобразуем в DataFrame
        df = pd.DataFrame(results, columns=["key", "value"])
        return df

    def update_carriage_info(self):
        if self.carriage_info_updated:
            df = self.extract_widget_data(self.carriage_info_formLayout)
            # print(df)
            for index, row in df.iterrows():
                # print(row['value'], type(row['value']))
                row['key'] = row['key'].replace('_text','')
                if isinstance(row['value'], bool):
                    row['value'] = str(int(row['value']))
                elif isinstance(row['value'], str):
                    row['value'] = (None if row['value'] == '' else row['value'].strip())
                # print(row['value'], type(row['value']))
            # print(df)
            df_transposed = df.set_index('key').T
            # print(df_transposed)
            df_transposed.to_sql('tmp_carriage_info_update', self.sqlite_conn, if_exists='replace', index=False)
            self.carriage_info_updated = False
            self.sqlite_cursor.execute("""SELECT GROUP_CONCAT(name||'=t2.'||name,',') list FROM PRAGMA_TABLE_INFO('tmp_carriage_info_update');""")
            rows = self.sqlite_cursor.fetchall()
            self.sqlite_cursor.execute(f"""UPDATE carriage SET {rows[0][0]} FROM tmp_carriage_info_update AS t2 WHERE carriage.carriage_code = t2.carriage_code;""")
            self.sqlite_cursor.execute("COMMIT")
            self.carriage_info(self.carriage_input.text())
            self._logging_message(f"[{self.info_mark}] Информация о вагоне обновлена")

    def update_factory_info(self):
        if self.factory_name_edit.text():
            new_name = self.factory_name_edit.text().strip()
            f_code = int(self.factory_input.text().strip())
            self.sqlite_cursor.execute(f"UPDATE factory SET factory_name = '{new_name}' where cast(factory_code as integer) = {f_code}")
            self.sqlite_cursor.execute("COMMIT")
            self.show_message("Информация", "Операция выполнена")

    def replacement_analysis(self):
        # noinspection PyUnresolvedReferences
        state = self.checkBox_full_analysis.checkState() == Qt.Checked
        self.sa = sa(self.sqlite_conn, self.sqlite_cursor, state)
        self.sa.sqlite_analysis_process()
        self.sa.sqlite_analysis_process2()
        self.checkBox_full_analysis.setChecked(False)
        # self.replecement_analysis_test()  # test
        self._logging_message(f"[{self.info_mark}] Анализ данных завершён")
        self._logging_message(f"[{self.info_mark}] {datetime.now().strftime('%H:%M:%S')}")
        self._logging_message("=" * 40)
        self._logging_message(" " * 40)

    # def replecement_analysis_test(self):
    #     self.show_message("Информация", "тест запускается")
    #     for i in ['export_repairs', 'repairs', 'repairs_copy']:  #
    #         # SQL-запрос (любой, например, результат анализа)
    #         query = f"SELECT * FROM {i}"
    #         # Загружаем результат запроса в DataFrame
    #         df = pd.read_sql_query(query, self.sqlite_conn)
    #         # Экспортируем в Excel
    #         df.to_excel(f"{i}.xlsx", index=False, engine='openpyxl')
    #         print(f"Результат успешно экспортирован в {i}.xlsx")
    #     self.show_message("Информация", "тест выполнен")

    def export_reports(self):

        report_folder = os.path.join(self.folder_input.text(), "Reports")

        os.makedirs(report_folder, exist_ok=True)
        now = datetime.now()
        template_wb = load_workbook("bin/Вагоны.xltx")
        template_wb.template = False
        report_file_path = f"{report_folder}/Вагоны_{now.strftime('%Y-%m-%d_%H.%M.%S')}.xlsx"
        template_wb.save(report_file_path)

        self.si.export_process()

        for f in [("export_repairs","Ремонты", 3),
                  ("export_equipment", "Замена деталей", 2),
                  ("export_equipment_details", "Последний ремонт", 2),
                  ("export_axis_analysis_full", "Срок службы", 2),
                  ("export_axis_analysis_final", "Срок службы итог", 2)]:

            export_file = pd.read_sql(f"SELECT * FROM {f[0]};", self.sqlite_conn)
            export_file['repair_date_iso'] = pd.to_datetime(export_file['repair_date_iso'], format='%Y-%m-%d').dt.date

            ##################################################################
            # # debug for Power BI
            # # dataframe to temp xlsx files - not used anymore
            # os.makedirs('Temp', exist_ok=True)
            # export_file_path = f"Temp/{f[0]}.xlsx"
            # export_file.to_excel(export_file_path, index=False)
            # wb = load_workbook(export_file_path)
            # ws = wb.active  # или wb['ИмяЛиста'], если знаешь имя листа
            # # Идём по столбцу A, начиная со второй строки
            # row = 2
            # while True:
            #     cell = ws[f"A{row}"]
            #     cellC = ws[f"C{row}"]  # check for all reports
            #     if cellC.value is None:
            #         break  # как только пустая ячейка — выходим
            #     cell.number_format = 'DD-MMM-YY'
            #     row += 1
            # ws.freeze_panes = "A2"
            # wb.save(export_file_path)
            ##################################################################

            # directly to template
            sheet_name = f[1]
            start_row = f[2]  # С какой строки вставлять (1-индексация)
            start_col = 1  # С какого столбца (например, 3 = колонка 'C')

            # Загружаем книгу и лист
            wb = load_workbook(report_file_path)
            ws = wb[sheet_name]

            # Вставляем данные вручную по ячейкам + исправляем форматы
            for i, row in export_file.iterrows():
                excel_row = start_row + i
                ws[f"A{excel_row}"].number_format = 'DD-MMM-YY'  # repair_date_iso = column A
                for j, value in enumerate(row):
                    cell = ws.cell(row=excel_row, column=start_col + j)
                    cell.value = value
                    cell.font = Font(name='Calibri', size=11)

            # # Также вставим заголовки => start_row+1
            # for j, colname in enumerate(export_file.columns):
            #     ws.cell(row=start_row - 1, column=start_col + j, value=colname)

            # Сохраняем файл
            wb.save(report_file_path)

            report_file_path_str = re.sub(r'\\', '/', report_file_path)

        self.sqlite_conn.execute("VACUUM")
        self._logging_message(f"[{self.info_mark}] Отчёт подготовлен в {report_file_path_str}")
        self._logging_message(f"[{self.info_mark}] {datetime.now().strftime('%H:%M:%S')}")
        self._logging_message("=" * 40)
        self._logging_message(" " * 40)
        self.show_message("Информация", "Операция выполнена")

        # folder_path = report_file_path_str
        # QDesktopServices.openUrl(QUrl.fromLocalFile(folder_path))

        # file_path = report_file_path_str
        # folder_path, file_name = os.path.split(file_path)
        # print(folder_path, file_name)
        # url = QUrl.fromLocalFile(folder_path)
        # url.setUrl(f"file:///{folder_path}/?select={file_name}")
        # QDesktopServices.openUrl(url)

        report_folder_path = os.path.dirname(report_file_path_str)
        # Открываем папку с выделенным файлом
        subprocess.run(["explorer", "/select,", os.path.normpath(report_file_path_str)])
        # Делаем небольшую паузу, чтобы Explorer успел запуститься
        time.sleep(0.5)
        # Находим окно с заголовком папки
        window_title = os.path.basename(report_folder_path)
        hwnd = win32gui.FindWindow(None, window_title)
        # Перемещаем окно на передний план
        if hwnd:
            win32gui.SetForegroundWindow(hwnd)
            win32gui.ShowWindow(hwnd, win32con.SW_SHOWNORMAL)

        if self.current_base_type_key == 'empty':
            self._logging_message(f"[{self.test_mark}] Нужна очистка этой базы? Пока не реализована")
            self._logging_message("")


class FolderOnlyPicker(QDialog):
    def __init__(self, parent=None, title="Выбор папки"):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(600, 400)

        # Устанавливаем иконку окна диалога
        self.setWindowIcon(QIcon(app_icon_png))

        layout = QVBoxLayout(self)

        # Строка с адресом выбранной папки
        self.path_line = QLineEdit()
        self.path_line.setReadOnly(True)
        # self.path_line.returnPressed.connect(self.handle_manual_path_input)
        layout.addWidget(self.path_line)

        # Модель для файловой системы
        self.model = QFileSystemModel()
        self.model.setRootPath(QDir.rootPath())
        # self.model.setFilter(QDir.AllDirs | QDir.NoDotAndDotDot)  # Только папки
        self.model.setFilter(QDir.AllEntries | QDir.NoDotAndDotDot)  # Папки и файлы

        # Виджет для дерева
        self.tree = QTreeView()
        self.tree.setModel(self.model)
        self.tree.setRootIndex(self.model.index(QDir.rootPath()))
        self.tree.setSelectionMode(QTreeView.SingleSelection)  # Выбор только одной папки
        self.tree.setColumnWidth(0, 400)
        # self.tree.setItemDelegateForColumn(0, self.IconDelegate())  # Устанавливаем иконки для папок и файлов --?
        layout.addWidget(self.tree)

        self.buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.buttons.accepted.connect(self.check_selection)
        self.buttons.rejected.connect(self.reject)
        layout.addWidget(self.buttons)

        # Обработчик изменения выбора
        self.tree.selectionModel().currentChanged.connect(self.update_path_line)

    def update_path_line(self, index):
        self.path_line.setText(self.model.filePath(index))

    def check_selection(self):
        index = self.tree.currentIndex()
        if self.model.isDir(index):
            self.accept()
        else:
            QMessageBox.warning(self, "Ошибка", "Выберите, пожалуйста, папку.")

    def get_selected_folder(self):
        index = self.tree.currentIndex()
        return self.model.filePath(index)

    def set_default_path(self, path):
        if not os.path.exists(path):
            return

        # Нормализуем путь и разбиваем на части
        path = os.path.normpath(path)
        parts = path.split(os.sep)
        current_path = os.path.abspath(os.sep)  # начинается с корня (например, "C:\\" или "/")

        def expand_step(i):
            nonlocal current_path
            if i >= len(parts):
                index = self.model.index(current_path)
                self.tree.setCurrentIndex(index)
                self.tree.scrollTo(index, QTreeView.PositionAtCenter)
                self.update_path_line(index)
                return

            current_path = os.path.join(current_path, parts[i])
            index = self.model.index(current_path)

            if not index.isValid():
                QTimer.singleShot(100, lambda: expand_step(i))  # если индекс ещё не готов — ждём
                return

            self.tree.expand(index)

            # Делаем паузу перед следующим шагом, чтобы дать GUI отрисоваться
            QTimer.singleShot(100, lambda: expand_step(i + 1))

        expand_step(0)

    # # Обработчик для ручного ввода пути
    # def handle_manual_path_input(self):
    #     path = self.path_line.text().strip()
    #     if os.path.isdir(path):
    #         self.set_default_path(path)
    #     else:
    #         QMessageBox.warning(self, "Ошибка", "Указанный путь не существует или не является папкой.")
    #
    # # Класс для отображения иконок в дереве
    # class IconDelegate(QStyledItemDelegate):
    #     def paint(self, painter, option, index):
    #         if index.isValid():
    #             icon = index.model().fileIcon(index)
    #             if icon.isNull():
    #                 icon = QIcon.fromTheme("folder")  # дефолтная иконка папки
    #             option.icon = icon
    #         super().paint(painter, option, index)


##################################################################


def main():
    app = QApplication()
    myappid = 'asvd.repairs.python.01'  # for icon in taskbar
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
    window = XLSXLoader()
    window.setWindowTitle("АСВД ремонты")
    window.resize(900, 800)
    window.show()
    app.exec()


if __name__ == "__main__":
    main()
